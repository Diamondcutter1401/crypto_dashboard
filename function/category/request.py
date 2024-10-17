import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

url = "https://api.coingecko.com/api/v3/coins/categories"

headers = {"accept": "application/json"}

response = requests.get(url, headers=headers)

data = response.json()
category_data =[]

for category in data:
    category_data.append({
                'Name': category.get('name'),
                'Market Cap': category.get('market_cap'),
                'MKC change (24h)': category.get('market_cap_change_24h'),
                'Volume (24h)': category.get('volume_24h')
            })

        # Convert to DataFrame
df = pd.DataFrame(category_data)
df = df[~df['Name'].str.contains('Portfolio', na=False)]
df = df[~df['Name'].str.contains('Stablecoin', na=False)]
df = df[~df['Name'].str.contains('Holding', na=False)]

# Save data to a CSV file
df.to_csv("data/crypto_category_data.csv", index=False)
df.to_excel("data/crypto_category_data.xlsx", index=False)

workbook = load_workbook('data/crypto_category_data.xlsx')
sheet = workbook.active

# Create a NamedStyle for the comma formatting (for Column B)
comma_style = NamedStyle(name="comma_style")
comma_style.number_format = '#,##0.00'  # Comma style with two decimal places

# Create another NamedStyle for increased decimal places (for Column C)
decimal_style = NamedStyle(name="decimal_style")
decimal_style.number_format = '0.0000'  # Four decimal places

# Apply the comma style to Column B
for row in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
    cell = f'B{row}'  # Change 'B' to the column you need for comma style
    sheet[cell].style = comma_style

# Apply the decimal style to Column C
for row in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
    cell = f'C{row}'  # Change 'C' to the column you need for increased decimals
    sheet[cell].style = decimal_style

for row in range(2, sheet.max_row + 1):  # Assuming data starts from row 2
    cell = f'D{row}'  # Change 'B' to the column you need for comma style
    sheet[cell].style = comma_style

# Save the workbook with the new formatting
workbook.save('data/crypto_category_data.xlsx')

print("Data has been saved to crypto_category_data.csv")